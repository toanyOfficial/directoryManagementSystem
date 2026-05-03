from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel_schema import EXCEL_HEADERS

_COLUMN_WIDTHS: dict[str, float] = {
    "A": 20,
    "B": 20,
    "C": 20,
    "D": 28,
    "E": 18,
    "F": 36,
    "G": 44,
}
_MAX_EXCEL_ROWS = 1048576


@dataclass(frozen=True)
class ExcelCreationResult:
    success: bool
    message: str
    path: Path | None = None


class ExcelInitializer:
    """Create the stage-1 Excel master file template."""

    def __init__(self, default_filename: str = "directory_master.xlsx") -> None:
        self.default_filename = default_filename

    def create_template(self, directory: Path) -> ExcelCreationResult:
        return self.create_template_at(directory / self.default_filename)

    def create_template_at(self, target_path: Path) -> ExcelCreationResult:
        try:
            resolved_target_path = target_path.expanduser().resolve()
        except OSError as exc:
            return ExcelCreationResult(False, f"경로를 해석할 수 없습니다: {exc}")

        target_directory = resolved_target_path.parent
        if not target_directory.exists():
            return ExcelCreationResult(False, f"대상 경로가 존재하지 않습니다: {target_directory}")

        if not target_directory.is_dir():
            return ExcelCreationResult(False, f"대상 경로가 폴더가 아닙니다: {target_directory}")

        if resolved_target_path.exists():
            return ExcelCreationResult(False, f"이미 파일이 존재합니다: {resolved_target_path}", resolved_target_path)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "원장"

        self._write_headers(worksheet)
        self._apply_column_widths(worksheet)
        self._apply_data_validation(worksheet)
        self._apply_skip_column_visibility_rule(worksheet)
        self._apply_relative_path_formula(worksheet)
        worksheet.freeze_panes = "A2"

        try:
            workbook.save(resolved_target_path)
        except OSError as exc:
            return ExcelCreationResult(False, f"엑셀 파일 저장에 실패했습니다: {exc}")

        return ExcelCreationResult(True, f"엑셀 파일을 생성했습니다: {resolved_target_path}", resolved_target_path)

    def _write_headers(self, worksheet: Worksheet) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for column_index, header in enumerate(EXCEL_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _apply_column_widths(self, worksheet: Worksheet) -> None:
        for column_letter, width in _COLUMN_WIDTHS.items():
            worksheet.column_dimensions[column_letter].width = width

    def _apply_data_validation(self, worksheet: Worksheet) -> None:
        # A~D 열은 기존과 동일하게 custom 수식 기반 validation을 유지한다.
        for column_letter in ("A", "B", "C", "D"):
            validation = DataValidation(
                type="custom",
                formula1=(
                    f'=AND({column_letter}2<>"",'
                    f'EXACT({column_letter}2,LOWER({column_letter}2)),'
                    f'ISERROR(SEARCH(" ",{column_letter}2)),'
                    f'RIGHT({column_letter}2,1)<>".")'
                ),
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="입력 제한",
                error=(
                    "영문은 소문자로 입력하고, 공백 없이 입력해야 하며, 마지막 글자에 '.'을 사용할 수 없습니다. "
                    "최종 유효성 검사는 프로그램 기준을 따릅니다."
                ),
                promptTitle="입력 규칙",
                prompt="숫자/한글/영문 소문자/언더스코어/점만 사용하고 공백은 입력하지 마세요.",
            )
            validation.add(f"{column_letter}2:{column_letter}{_MAX_EXCEL_ROWS}")
            worksheet.add_data_validation(validation)

        # E 열은 별도의 validation 객체로 TRUE/FALSE 목록만 허용한다.
        e_validation = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="입력 제한",
            error="TRUE 또는 FALSE만 입력 가능합니다.",
            promptTitle="입력 규칙",
            prompt="TRUE(하위폴더 무시), FALSE(검사)",
        )
        e_validation.add(f"E2:E{_MAX_EXCEL_ROWS}")
        worksheet.add_data_validation(e_validation)

    def _apply_skip_column_visibility_rule(self, worksheet: Worksheet) -> None:
        # E열에서 FALSE 값은 삭제하지 않고 "보이지 않게" 처리하기 위해 조건부 서식을 적용한다.
        # EXACT($E2,"FALSE")가 참일 때 글자색을 흰색으로 바꿔 기본 배경(흰색)에서 숨김 효과를 낸다.
        false_hidden_rule = FormulaRule(
            formula=['EXACT($E2,"FALSE")'],
            font=Font(color="FFFFFF"),
        )
        worksheet.conditional_formatting.add(f"E2:E{_MAX_EXCEL_ROWS}", false_hidden_rule)

    def _apply_relative_path_formula(self, worksheet: Worksheet) -> None:
        # 템플릿 생성 시점부터 G열(상대경로)에 동일 수식을 선언해 즉시 사용할 수 있게 한다.
        worksheet["G2"] = '=TEXTJOIN("\\",TRUE,A2:D2)'
