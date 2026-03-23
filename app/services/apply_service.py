from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
import os
from pathlib import Path, PureWindowsPath
from shutil import copy2

from openpyxl import load_workbook

from app.services.dry_run_analyzer import DryRunAnalyzer, DryRunResult, ParsedRow


@dataclass(frozen=True)
class ApplyResult:
    success: bool
    message: str
    status_message: str
    target_root: Path | None = None
    backup_path: Path | None = None
    log_path: Path | None = None
    created_folders: list[str] = field(default_factory=list)
    deleted_folders: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    rollback_actions: list[str] = field(default_factory=list)
    hyperlink_updated_rows: int = 0
    dry_run_result: DryRunResult | None = None


class ApplyService:
    """Apply the validated directory changes and update Excel hyperlinks safely."""

    def __init__(self, dry_run_analyzer: DryRunAnalyzer) -> None:
        self.dry_run_analyzer = dry_run_analyzer

    def apply(self, excel_path: Path, root_directory: Path | None = None) -> ApplyResult:
        started_at = datetime.now()
        timestamp = started_at.strftime("%Y%m%d_%H%M%S")

        try:
            resolved_excel_path = excel_path.expanduser().resolve()
        except OSError as exc:
            return ApplyResult(
                success=False,
                message=f"엑셀 경로를 확인할 수 없습니다: {exc}",
                status_message="적용 실패",
                errors=[f"엑셀 경로를 확인할 수 없습니다: {exc}"],
            )

        if root_directory is None:
            target_root = resolved_excel_path.parent
        else:
            try:
                target_root = root_directory.expanduser().resolve()
            except OSError as exc:
                return ApplyResult(
                    success=False,
                    message=f"루트 디렉토리를 확인할 수 없습니다: {exc}",
                    status_message="적용 실패",
                    errors=[f"루트 디렉토리를 확인할 수 없습니다: {exc}"],
                )

            if not target_root.exists():
                return ApplyResult(
                    success=False,
                    message=f"선택한 루트 디렉토리가 존재하지 않습니다: {target_root}",
                    status_message="적용 실패",
                    errors=[f"선택한 루트 디렉토리가 존재하지 않습니다: {target_root}"],
                )

            if not target_root.is_dir():
                return ApplyResult(
                    success=False,
                    message=f"선택한 루트 경로가 폴더가 아닙니다: {target_root}",
                    status_message="적용 실패",
                    errors=[f"선택한 루트 경로가 폴더가 아닙니다: {target_root}"],
                )

        logs_directory = target_root / "logs"
        backups_directory = target_root / "backups"
        log_path = logs_directory / f"apply_{timestamp}.log"

        created_relative_paths: list[Path] = []
        deleted_relative_paths: list[Path] = []
        errors: list[str] = []
        rollback_actions: list[str] = []
        backup_path: Path | None = None
        dry_run_result: DryRunResult | None = None

        try:
            logs_directory.mkdir(parents=True, exist_ok=True)
            backups_directory.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            return ApplyResult(
                success=False,
                message=f"로그 또는 백업 폴더를 준비할 수 없습니다. 접근 권한을 확인해 주세요. ({exc})",
                status_message="적용 실패",
                target_root=target_root,
                log_path=log_path,
                errors=[f"로그 또는 백업 폴더를 준비할 수 없습니다: {exc}"],
            )

        try:
            dry_run_result = self.dry_run_analyzer.analyze(resolved_excel_path, target_root)
            if not dry_run_result.success:
                raise RuntimeError(dry_run_result.fatal_error or "사전 검증에 실패했습니다.")

            if not dry_run_result.is_applicable:
                raise RuntimeError("사전 검증 결과 적용 불가 상태입니다. row 오류 또는 위험 폴더를 먼저 해결하세요.")

            if dry_run_result.danger_relative_paths:
                raise RuntimeError("위험 폴더가 존재하여 적용을 중단합니다.")

            self._ensure_empty_delete_candidates(target_root, dry_run_result.delete_relative_paths)
            backup_path = backups_directory / f"{resolved_excel_path.stem}_{timestamp}{resolved_excel_path.suffix}"
            copy2(resolved_excel_path, backup_path)

            self._create_directories(target_root, dry_run_result.create_relative_paths, created_relative_paths)
            hyperlink_updated_rows = self._update_hyperlinks(
                resolved_excel_path,
                target_root,
                dry_run_result.parsed_rows,
            )
            self._ensure_empty_delete_candidates(target_root, dry_run_result.delete_relative_paths)
            self._delete_empty_directories(target_root, dry_run_result.delete_relative_paths, deleted_relative_paths)

            post_apply_result = self.dry_run_analyzer.analyze(resolved_excel_path, target_root)
            result = ApplyResult(
                success=True,
                message="적용이 완료되었습니다.",
                status_message="적용 완료",
                target_root=target_root,
                backup_path=backup_path,
                log_path=log_path,
                created_folders=[self._format_path(path) for path in created_relative_paths],
                deleted_folders=[self._format_path(path) for path in deleted_relative_paths],
                hyperlink_updated_rows=hyperlink_updated_rows,
                dry_run_result=post_apply_result,
            )
        except Exception as exc:
            errors.append(str(exc))
            rollback_actions, rollback_errors = self._rollback(
                excel_path=resolved_excel_path,
                backup_path=backup_path,
                target_root=target_root,
                created_relative_paths=created_relative_paths,
                deleted_relative_paths=deleted_relative_paths,
            )
            errors.extend(rollback_errors)
            result = ApplyResult(
                success=False,
                message="적용에 실패했습니다.",
                status_message="적용 실패",
                target_root=target_root,
                backup_path=backup_path,
                log_path=log_path,
                created_folders=[self._format_path(path) for path in created_relative_paths],
                deleted_folders=[self._format_path(path) for path in deleted_relative_paths],
                errors=errors,
                rollback_actions=rollback_actions,
                dry_run_result=dry_run_result,
            )

        log_errors = self._write_log(
            log_path=log_path,
            started_at=started_at,
            result=result,
        )

        if log_errors:
            return ApplyResult(
                success=result.success,
                message=f"{result.message} (로그 저장 경고)",
                status_message=result.status_message,
                target_root=result.target_root,
                backup_path=result.backup_path,
                log_path=result.log_path,
                created_folders=result.created_folders,
                deleted_folders=result.deleted_folders,
                errors=[*result.errors, *log_errors],
                rollback_actions=result.rollback_actions,
                hyperlink_updated_rows=result.hyperlink_updated_rows,
                dry_run_result=result.dry_run_result,
            )

        return result

    def _create_directories(
        self,
        target_root: Path,
        relative_paths: list[Path],
        created_relative_paths: list[Path],
    ) -> None:
        for relative_path in relative_paths:
            absolute_path = target_root / relative_path
            if absolute_path.exists():
                continue
            absolute_path.mkdir(parents=True, exist_ok=False)
            created_relative_paths.append(relative_path)

    def _update_hyperlinks(self, excel_path: Path, target_root: Path, parsed_rows: list[ParsedRow]) -> int:
        workbook = load_workbook(excel_path)
        try:
            worksheet = workbook.active
            for parsed_row in parsed_rows:
                cell = worksheet.cell(row=parsed_row.row_number, column=4)
                hyperlink_target = os.path.relpath(target_root / parsed_row.relative_path, start=excel_path.parent)
                cell.hyperlink = str(PureWindowsPath(hyperlink_target))
                cell.style = "Hyperlink"
            workbook.save(excel_path)
        finally:
            workbook.close()
        return len(parsed_rows)

    def _ensure_empty_delete_candidates(self, target_root: Path, relative_paths: list[Path]) -> None:
        for relative_path in relative_paths:
            absolute_path = target_root / relative_path
            try:
                has_children = any(absolute_path.iterdir())
            except OSError as exc:
                raise RuntimeError(
                    f"삭제 대상 폴더 접근에 실패했습니다: {self._format_path(relative_path)} ({exc})"
                ) from exc
            if has_children:
                raise RuntimeError(f"삭제 대상 폴더가 비어 있지 않습니다: {self._format_path(relative_path)}")

    def _delete_empty_directories(
        self,
        target_root: Path,
        relative_paths: list[Path],
        deleted_relative_paths: list[Path],
    ) -> None:
        for relative_path in sorted(relative_paths, key=lambda path: (len(path.parts), self._format_path(path)), reverse=True):
            absolute_path = target_root / relative_path
            absolute_path.rmdir()
            deleted_relative_paths.append(relative_path)

    def _rollback(
        self,
        excel_path: Path,
        backup_path: Path | None,
        target_root: Path,
        created_relative_paths: list[Path],
        deleted_relative_paths: list[Path],
    ) -> tuple[list[str], list[str]]:
        rollback_actions: list[str] = []
        rollback_errors: list[str] = []

        if backup_path is not None and backup_path.exists():
            try:
                copy2(backup_path, excel_path)
                rollback_actions.append(f"엑셀 백업을 원본으로 복원했습니다: {backup_path}")
            except OSError as exc:
                rollback_errors.append(f"엑셀 복원 실패: {exc}")

        for relative_path in sorted(deleted_relative_paths, key=lambda path: (len(path.parts), self._format_path(path))):
            try:
                (target_root / relative_path).mkdir(parents=True, exist_ok=True)
                rollback_actions.append(f"삭제한 빈 폴더를 다시 만들었습니다: {self._format_path(relative_path)}")
            except OSError as exc:
                rollback_errors.append(f"삭제 폴더 복원 실패 ({self._format_path(relative_path)}): {exc}")

        for relative_path in sorted(
            created_relative_paths,
            key=lambda path: (len(path.parts), self._format_path(path)),
            reverse=True,
        ):
            absolute_path = target_root / relative_path
            try:
                if absolute_path.exists() and not any(absolute_path.iterdir()):
                    absolute_path.rmdir()
                    rollback_actions.append(f"생성했던 폴더를 제거했습니다: {self._format_path(relative_path)}")
            except OSError as exc:
                rollback_errors.append(f"생성 폴더 롤백 실패 ({self._format_path(relative_path)}): {exc}")

        return rollback_actions, rollback_errors

    def _write_log(self, log_path: Path, started_at: datetime, result: ApplyResult) -> list[str]:
        ended_at = datetime.now()
        lines = [
            f"실행 시간: {started_at.strftime('%Y-%m-%d %H:%M:%S')}",
            f"종료 시간: {ended_at.strftime('%Y-%m-%d %H:%M:%S')}",
            f"대상 루트: {result.target_root if result.target_root else '-'}",
            f"백업 파일: {result.backup_path if result.backup_path else '-'}",
            "생성 폴더:",
            *self._format_log_items(result.created_folders),
            "삭제 폴더:",
            *self._format_log_items(result.deleted_folders),
            f"갱신된 하이퍼링크 row 수: {result.hyperlink_updated_rows}",
            "오류:",
            *self._format_log_items(result.errors),
            "롤백 시도:",
            *self._format_log_items(result.rollback_actions),
            f"결과: {result.message}",
        ]

        try:
            log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        except OSError as exc:
            return [f"로그 파일 저장 실패: {exc}"]

        return []

    def _format_log_items(self, items: list[str]) -> list[str]:
        if not items:
            return ["- 없음"]
        return [f"- {item}" for item in items]

    def _format_path(self, path: Path) -> str:
        return str(PureWindowsPath(*path.parts))
