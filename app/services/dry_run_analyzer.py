from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path, PureWindowsPath

from openpyxl import load_workbook

from app.services.excel_schema import EXCEL_HEADERS
from app.utils.path_validator import FolderNameValidator

_SYSTEM_DIRECTORIES = {"logs", "backups", "_internal"}


@dataclass(frozen=True)
class RowError:
    row_number: int
    message: str


@dataclass(frozen=True)
class ParsedRow:
    row_number: int
    path_parts: tuple[str, ...]
    hyperlink_column: int

    @property
    def relative_path(self) -> Path:
        return Path(*self.path_parts)

    @property
    def display_path(self) -> str:
        return str(PureWindowsPath(*self.path_parts))


@dataclass(frozen=True)
class DryRunResult:
    success: bool
    fatal_error: str | None
    target_root: Path | None
    total_rows: int
    valid_rows: int
    error_rows: int
    create_count: int
    delete_count: int
    review_needed_count: int
    is_applicable: bool
    create_candidates: list[str] = field(default_factory=list)
    delete_candidates: list[str] = field(default_factory=list)
    review_needed_items: list[str] = field(default_factory=list)
    row_errors: list[RowError] = field(default_factory=list)
    parsed_rows: list[ParsedRow] = field(default_factory=list)
    create_relative_paths: list[Path] = field(default_factory=list)
    delete_relative_paths: list[Path] = field(default_factory=list)
    review_needed_relative_paths: list[Path] = field(default_factory=list)


class DryRunAnalyzer:
    """Analyze Excel rows and compare them with the actual directory structure."""

    def analyze(self, excel_path: Path, root_directory: Path | None = None) -> DryRunResult:
        try:
            resolved_excel_path = excel_path.expanduser().resolve()
        except OSError as exc:
            return self._fatal_result(f"엑셀 경로를 확인할 수 없습니다. 경로를 다시 선택해 주세요. ({exc})")

        if not resolved_excel_path.exists():
            return self._fatal_result(f"선택한 엑셀 파일을 찾을 수 없습니다: {resolved_excel_path}")

        if resolved_excel_path.suffix.lower() != ".xlsx":
            return self._fatal_result("지원하지 않는 파일 형식입니다. '.xlsx' 파일을 선택해 주세요.")

        if root_directory is None:
            target_root = resolved_excel_path.parent
        else:
            try:
                target_root = root_directory.expanduser().resolve()
            except OSError as exc:
                return self._fatal_result(f"루트 디렉토리를 확인할 수 없습니다. 경로를 다시 선택해 주세요. ({exc})")

            if not target_root.exists():
                return self._fatal_result(f"선택한 루트 디렉토리가 존재하지 않습니다: {target_root}")

            if not target_root.is_dir():
                return self._fatal_result(f"선택한 루트 경로가 폴더가 아닙니다: {target_root}")

        try:
            workbook = load_workbook(resolved_excel_path, data_only=True)
        except Exception as exc:
            return self._fatal_result(f"엑셀 파일을 읽는 중 문제가 발생했습니다. 파일이 열려 있는지 확인해 주세요. ({exc})")

        try:
            worksheet = workbook.active
            header_values = tuple(
                str(worksheet.cell(row=1, column=index).value or "").strip() for index in range(1, 6)
            )
            if header_values != EXCEL_HEADERS:
                return self._fatal_result(
                    "엑셀 헤더가 올바르지 않습니다. "
                    f"기대값: {', '.join(EXCEL_HEADERS)} / 실제값: {', '.join(header_values)}"
                )

            row_errors: list[RowError] = []
            parsed_rows: list[ParsedRow] = []
            seen_paths: set[Path] = set()
            total_rows = 0

            depth_column_count = len(EXCEL_HEADERS) - 1
            for row_number in range(2, worksheet.max_row + 1):
                values = [worksheet.cell(row=row_number, column=index).value for index in range(1, len(EXCEL_HEADERS) + 1)]
                normalized_values: list[str] = []
                for index, value in enumerate(values):
                    if value is None:
                        normalized_values.append("")
                    elif index < depth_column_count:
                        normalized_values.append(FolderNameValidator.normalize(str(value)))
                    else:
                        normalized_values.append(str(value).strip())

                if not any(normalized_values):
                    continue

                total_rows += 1
                validation_messages = self._validate_row(normalized_values, depth_column_count)
                if validation_messages:
                    row_errors.append(RowError(row_number, "; ".join(validation_messages)))
                    continue

                depth_values = normalized_values[:depth_column_count]
                path_parts = tuple(value for value in depth_values if value)
                hyperlink_column = self._last_value_index(depth_values) + 1
                parsed_row = ParsedRow(
                    row_number=row_number,
                    path_parts=path_parts,
                    hyperlink_column=hyperlink_column,
                )
                if parsed_row.relative_path in seen_paths:
                    row_errors.append(RowError(row_number, f"중복 구조입니다: {parsed_row.display_path}"))
                    continue

                seen_paths.add(parsed_row.relative_path)
                parsed_rows.append(parsed_row)
        finally:
            workbook.close()

        expected_directories = self._build_expected_directories(parsed_rows)
        depth_column_count = len(EXCEL_HEADERS) - 1
        actual_directories = self._scan_actual_directories(target_root, depth_column_count)

        create_candidates = sorted(expected_directories - actual_directories, key=self._sort_key)
        leaf_directories = {parsed_row.relative_path for parsed_row in parsed_rows}

        # 삭제 후보 판단 로직
        # 1) expected에 없는 실제 폴더
        # 2) 관리 depth(현재 Depth1~Depth4) 이내
        # 3) 부모가 leaf_directories인 경우 제외 (leaf 아래 데이터 구조 보호)
        delete_candidates = sorted(
            self._build_delete_candidates(
                actual_directories=actual_directories,
                expected_directories=expected_directories,
                leaf_directories=leaf_directories,
                managed_depth=depth_column_count,
            ),
            key=self._sort_key,
        )

        # 확인필요 판단 로직
        # - 엑셀 leaf 경로가 실제로 존재하고
        # - leaf 바로 아래 하위 폴더가 1~5개인 경우만 표시
        review_needed = self._build_review_needed_items(
            target_root=target_root,
            leaf_directories=leaf_directories,
            max_child_directory_count=5,
        )

        return DryRunResult(
            success=True,
            fatal_error=None,
            target_root=target_root,
            total_rows=total_rows,
            valid_rows=len(parsed_rows),
            error_rows=len(row_errors),
            create_count=len(create_candidates),
            delete_count=len(delete_candidates),
            review_needed_count=len(review_needed),
            # apply 차단 기준: row 오류 또는 삭제 후보 존재
            is_applicable=(len(row_errors) == 0 and len(delete_candidates) == 0),
            create_candidates=[self._format_path(path) for path in create_candidates],
            delete_candidates=[self._format_path(path) for path in delete_candidates],
            review_needed_items=[item["display"] for item in review_needed],
            row_errors=row_errors,
            parsed_rows=parsed_rows,
            create_relative_paths=create_candidates,
            delete_relative_paths=delete_candidates,
            review_needed_relative_paths=[item["path"] for item in review_needed],
        )

    def _validate_row(self, row_values: list[str], depth_column_count: int) -> list[str]:
        depth_values = row_values[:depth_column_count]
        errors: list[str] = []

        if not depth_values[0]:
            errors.append("Depth1은 필수입니다.")

        seen_empty = False
        for index, value in enumerate(depth_values, start=1):
            if not value:
                seen_empty = True
                continue
            if seen_empty:
                errors.append("중간 누락 구조는 허용되지 않습니다.")
                break

            field_name = f"Depth{index}"
            if not value:
                continue
            validation_result = FolderNameValidator.validate(value)
            if not validation_result.is_valid:
                errors.append(f"{field_name}: {validation_result.message}")

        return errors

    def _build_expected_directories(self, parsed_rows: list[ParsedRow]) -> set[Path]:
        expected_directories: set[Path] = set()
        for parsed_row in parsed_rows:
            current_path = Path()
            for part in parsed_row.path_parts:
                current_path /= part
                expected_directories.add(current_path)
        return expected_directories

    def _build_delete_candidates(
        self,
        actual_directories: set[Path],
        expected_directories: set[Path],
        leaf_directories: set[Path],
        managed_depth: int,
    ) -> set[Path]:
        delete_candidates: set[Path] = set()
        for actual_path in actual_directories:
            if actual_path in expected_directories:
                continue
            if len(actual_path.parts) > managed_depth:
                continue
            if actual_path.parent in leaf_directories:
                continue
            delete_candidates.add(actual_path)
        return delete_candidates

    def _build_review_needed_items(
        self,
        target_root: Path,
        leaf_directories: set[Path],
        max_child_directory_count: int,
    ) -> list[dict[str, object]]:
        review_needed: list[dict[str, object]] = []
        for leaf_path in sorted(leaf_directories, key=self._sort_key):
            absolute_leaf_path = target_root / leaf_path
            if not absolute_leaf_path.exists() or not absolute_leaf_path.is_dir():
                continue
            # leaf 바로 아래 하위 폴더만 집계한다.
            child_directories = self._list_child_directories(absolute_leaf_path)
            child_count = len(child_directories)
            if child_count == 0 or child_count > max_child_directory_count:
                continue
            sample_children = [child.name for child in child_directories[:max_child_directory_count]]
            display = (
                f"{self._format_path(leaf_path)} "
                f"(하위폴더 {child_count}개: {', '.join(sample_children)})"
            )
            review_needed.append(
                {
                    "path": leaf_path,
                    "display": display,
                }
            )
        return review_needed

    def _list_child_directories(self, directory_path: Path) -> list[Path]:
        try:
            return sorted([child for child in directory_path.iterdir() if child.is_dir()], key=lambda path: path.name)
        except OSError:
            return []

    def _scan_actual_directories(self, target_root: Path, max_depth: int) -> set[Path]:
        actual_directories: set[Path] = set()
        for path in target_root.rglob("*"):
            if not path.is_dir():
                continue
            relative_path = path.relative_to(target_root)
            if relative_path.parts and relative_path.parts[0] in _SYSTEM_DIRECTORIES:
                continue
            if len(relative_path.parts) > max_depth:
                continue
            actual_directories.add(relative_path)
        return actual_directories

    def _sort_key(self, path: Path) -> tuple[int, str]:
        return (len(path.parts), self._format_path(path))

    def _format_path(self, path: Path) -> str:
        return str(PureWindowsPath(*path.parts))

    def _last_value_index(self, values: list[str]) -> int:
        for index in range(len(values) - 1, -1, -1):
            if values[index]:
                return index
        return 0

    def _fatal_result(self, message: str) -> DryRunResult:
        return DryRunResult(
            success=False,
            fatal_error=message,
            target_root=None,
            total_rows=0,
            valid_rows=0,
            error_rows=0,
            create_count=0,
            delete_count=0,
            review_needed_count=0,
            is_applicable=False,
        )
